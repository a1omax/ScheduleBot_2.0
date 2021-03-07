from os import listdir
import openpyxl
from openpyxl.utils import rows_from_range
import time
FILENAME = "Розклад занять на 2 семестр 2020-2021 (з 01.03.2021).xlsx"

start_time = time.time()
def passed(start_time):
    time.time()
    print("Time passed: "+str(time.time()-start_time)+" seconds")

def check():
    if FILENAME in listdir():
        return True
    else:
        return False


def readWb(FILENAME,numberOfList):
    global wb, ws
    wb = openpyxl.load_workbook(filename=FILENAME)
    ws = wb[wb.sheetnames[numberOfList]]

def mergedCell(cell):
    idx = cell.coordinate
    for range_ in ws.merged_cells.ranges:
        merged_cells = list(openpyxl.utils.rows_from_range(str(range_)))
        for row in merged_cells:
            if idx in row:
                # If this is a merged cell,
                # return  the first cell of the merge range
                return ws[merged_cells[0][0]].value


def getValues(arr):

    valArr = []

    for i in arr:
        if str(type(i)) == "<class 'openpyxl.cell.cell.Cell'>":
            val = i.value
            if val == None:
                valArr.append("нет пары")
            else:
                valArr.append(i.value)
        elif str(type(i)) == "<class 'openpyxl.cell.cell.MergedCell'>":
            val = mergedCell(i)
            if val == None:
                valArr.append("нет пар")
            else:
                valArr.append(val)
    return valArr


def sched(col, start):
    letter = openpyxl.utils.cell.get_column_letter(col)
    arraySchedule = []
    for i in range(35):
        arraySchedule.append(ws[letter+str(start+i*4)])

    return getValues(arraySchedule)



def weekGenerateArray(numGroup):

    col = numGroup * 3
    firstSubGrOdd = sched(col, 6)
    secondSubGrOdd = sched(col + 1, 6)
    firstSubGrEven = sched(col, 8)
    secondSubGrEven = sched(col + 1, 8)
    oddWeek = []
    evenWeek = []
    for i in range(35):
        oddWeek.append([firstSubGrOdd[i], secondSubGrOdd[i]])
        evenWeek.append([firstSubGrEven[i], secondSubGrEven[i]])

    bothweeks = [oddWeek, evenWeek]

    return bothweeks
    #массив обеих подгрупп для чет и нечет недель

def groups(numGroup):
    col = numGroup * 3
    letter = openpyxl.utils.cell.get_column_letter(col)
    a = [ws[letter + str(5)]]
    getValues(a)
    return getValues(a)[0]

def generateSchedule(group):

    array = weekGenerateArray(group)
    schedule = []

    for d in range(2):
        dictTypeOfWeek = {}
        for day in range(7):
            dictDay = {}
            for para in range(1,6):
                dictDay[para] = (array[d][(para-1)+day*5])
            dictTypeOfWeek[day] = dictDay
        schedule.append(dictTypeOfWeek)
    return schedule



def main():
    start_time = time.time()
    print("Looking for a file")

    if check():

        #deleting data
        with open("schedules.py", "w", encoding="utf8") as f:
            f.write("")
        with open("groups.py", "w", encoding="utf8") as f:
            f.write("")


        print("File is found. Start parsing...")
        iter_ = [34, 28, 29, 28, 25, 19, 10, 20]
        dictGroup = {}
        listSched = []


        for j in range(len(iter_)):
            print("List number {} is now in progress".format(j))
            readWb(FILENAME, j)
            toWrite = []
            grouparr = []
            for i in range(1,iter_[j]):
                toWrite.append(generateSchedule(i))
                grouparr.append(str(groups(i)))


            passed(start_time)
            dictGroup[j] = grouparr
            listSched.append(toWrite)

        with open("schedules.py", "w", encoding="utf8") as f:
            f.write("\nsched = {write}".format(write=str(listSched)))

        with open("groups.py", "a", encoding="utf8") as f:
            f.write("group = {}".format(dictGroup))
        print("dict groups generated")
    else:
        print("File is not found")
        return -1

main()
